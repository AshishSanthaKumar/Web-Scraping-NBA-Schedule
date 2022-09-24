from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

#Defining the workbook and column headers 
wb = Workbook()
ws = wb.active

ws['A1'] = "Match Day"
ws['B1'] = "Match Time"
ws['C1'] = "Visitor Team"
ws['D1'] = "Home Team"
ws['E1'] = "Arena"

#Iterating through URL

months = ['october','november','december','january','february','march','april']
for month in months:
    
    url = f"https://www.basketball-reference.com/leagues/NBA_2023_games-{month}.html"
    response = requests.get(url)
    html = response.text
      
    #reading the contents of each URL
    soup = BeautifulSoup(html, "html.parser")
    
    Schedule = soup.find('tbody')
    
    for match in Schedule.find_all('tr'):
        match_day = match.find('th').text
        match_time = match.find('td').text
        list1 = [match_day, match_time]
        for td in match.find_all('td', class_='left'):
            list1.append(td.text)
            
        ws.append(list1)


# Save the file
wb.save("sample.xlsx")
